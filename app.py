import os
from datetime import datetime, date
from functools import wraps

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, make_response
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
import csv
import io

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_XLSX = os.path.join(BASE_DIR, "SHADMER.xlsx")
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///shop_data.db")
# Render Postgres URLs can start with postgres://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DATABASE_URL, future=True)
SessionLocal = sessionmaker(bind=engine, future=True)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "shadmer26-change-this-secret-key")


def money(value):
    try:
        return f"GHS {float(value):,.2f}"
    except Exception:
        return "GHS 0.00"


app.jinja_env.filters["money"] = money


def init_db():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                full_name TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL CHECK (role IN ('admin', 'attendant')),
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS inventory (
                id INTEGER PRIMARY KEY,
                product_name TEXT UNIQUE NOT NULL,
                category TEXT,
                cost_price REAL DEFAULT 0,
                selling_price REAL DEFAULT 0,
                stock_qty INTEGER DEFAULT 0,
                reorder_level INTEGER DEFAULT 0,
                expiry_date TEXT,
                created_at TEXT,
                updated_at TEXT
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY,
                sale_date TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                unit_price REAL NOT NULL,
                total_price REAL NOT NULL,
                payment_method TEXT,
                attendant_name TEXT,
                FOREIGN KEY(product_id) REFERENCES inventory(id)
            )
        """))
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS purchases (
                id INTEGER PRIMARY KEY,
                purchase_date TEXT NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                cost_price REAL NOT NULL,
                total_cost REAL NOT NULL,
                attendant_name TEXT,
                FOREIGN KEY(product_id) REFERENCES inventory(id)
            )
        """))

        existing_users = conn.execute(text("SELECT COUNT(*) FROM users")).scalar_one()
        if existing_users == 0:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn.execute(text("""
                INSERT INTO users (username, full_name, password_hash, role, created_at)
                VALUES
                (:u1, :n1, :p1, 'admin', :created_at),
                (:u2, :n2, :p2, 'attendant', :created_at)
            """), {
                "u1": "shadmer2026",
                "n1": "Shop Admin",
                "p1": generate_password_hash("shadmer002026"),
                "u2": "attendant",
                "n2": "Shop Attendant",
                "p2": generate_password_hash("attendant123"),
                "created_at": now,
            })

        inventory_count = conn.execute(text("SELECT COUNT(*) FROM inventory")).scalar_one()
        if inventory_count == 0 and os.path.exists(SOURCE_XLSX):
            import_inventory_from_excel(conn, SOURCE_XLSX)


def import_inventory_from_excel(conn, xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    if "Inventory" not in wb.sheetnames:
        return
    ws = wb["Inventory"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row[:7]) + [None] * max(0, 7 - len(row[:7]))
        product_name, category, cost_price, selling_price, stock_qty, reorder_level, expiry_date = values[:7]
        if not product_name:
            continue
        product_name = str(product_name).strip()
        expiry_text = ""
        if expiry_date:
            expiry_text = expiry_date.strftime("%Y-%m-%d") if hasattr(expiry_date, "strftime") else str(expiry_date)
        conn.execute(text("""
            INSERT OR IGNORE INTO inventory
            (product_name, category, cost_price, selling_price, stock_qty, reorder_level, expiry_date, created_at, updated_at)
            VALUES (:product_name, :category, :cost_price, :selling_price, :stock_qty, :reorder_level, :expiry_date, :created_at, :updated_at)
        """), {
            "product_name": product_name,
            "category": (str(category).strip() if category else "Beauty"),
            "cost_price": float(cost_price or 0),
            "selling_price": float(selling_price or 0),
            "stock_qty": int(stock_qty or 0),
            "reorder_level": int(reorder_level or 0),
            "expiry_date": expiry_text,
            "created_at": now,
            "updated_at": now,
        })


def require_login(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapped


def require_role(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login"))
            if session.get("role") not in roles:
                flash("You do not have access to that page.", "error")
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


def get_db():
    return SessionLocal()


def current_user_dict():
    return {
        "id": session.get("user_id"),
        "username": session.get("username"),
        "full_name": session.get("full_name"),
        "role": session.get("role"),
    }


@app.context_processor
def inject_globals():
    return {"current_user": current_user_dict(), "today": date.today().isoformat()}


@app.route("/")
def index():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        db = get_db()
        try:
            user = db.execute(text("SELECT * FROM users WHERE username = :username AND is_active = 1"), {"username": username}).mappings().first()
        finally:
            db.close()
        if not user or not check_password_hash(user["password_hash"], password):
            flash("Wrong username or password.", "error")
            return render_template("login.html")

        session["user_id"] = user["id"]
        session["username"] = user["username"]
        session["full_name"] = user["full_name"]
        session["role"] = user["role"]
        flash(f"Welcome, {user['full_name']}!", "success")
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/logout")
@require_login
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))


@app.route("/dashboard")
@require_login
def dashboard():
    db = get_db()
    try:
        kpis = db.execute(text("""
            SELECT
                COALESCE((SELECT SUM(total_price) FROM sales), 0) AS total_sales,
                COALESCE((SELECT SUM(quantity) FROM sales), 0) AS sold_units,
                COALESCE((SELECT SUM(stock_qty) FROM inventory), 0) AS stock_units,
                COALESCE((SELECT COUNT(*) FROM inventory), 0) AS products,
                COALESCE((SELECT SUM(total_price) FROM sales WHERE DATE(sale_date) = :today), 0) AS today_sales
        """), {"today": date.today().isoformat()}).mappings().first()

        low_stock = db.execute(text("""
            SELECT product_name, stock_qty, reorder_level
            FROM inventory
            WHERE stock_qty <= COALESCE(reorder_level, 0)
            ORDER BY stock_qty ASC, product_name ASC
            LIMIT 12
        """)).mappings().all()

        recent_sales = db.execute(text("""
            SELECT s.sale_date, i.product_name, s.quantity, s.total_price, s.payment_method, s.attendant_name
            FROM sales s
            JOIN inventory i ON i.id = s.product_id
            ORDER BY s.id DESC
            LIMIT 10
        """)).mappings().all()
    finally:
        db.close()
    return render_template("dashboard.html", kpis=kpis, low_stock=low_stock, recent_sales=recent_sales)


@app.route("/inventory", methods=["GET", "POST"])
@require_login
def inventory():
    db = get_db()
    try:
        if request.method == "POST":
            product_name = request.form.get("product_name", "").strip()
            category = request.form.get("category", "Beauty").strip()
            selling_price = float(request.form.get("selling_price", 0) or 0)
            cost_price = float(request.form.get("cost_price", 0) or 0)
            stock_qty = int(request.form.get("stock_qty", 0) or 0)
            reorder_level = int(request.form.get("reorder_level", 0) or 0)
            if not product_name:
                raise ValueError("Product name is required.")
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db.execute(text("""
                INSERT INTO inventory (product_name, category, cost_price, selling_price, stock_qty, reorder_level, expiry_date, created_at, updated_at)
                VALUES (:product_name, :category, :cost_price, :selling_price, :stock_qty, :reorder_level, '', :created_at, :updated_at)
            """), {
                "product_name": product_name,
                "category": category,
                "cost_price": cost_price,
                "selling_price": selling_price,
                "stock_qty": stock_qty,
                "reorder_level": reorder_level,
                "created_at": now,
                "updated_at": now,
            })
            db.commit()
            flash(f"{product_name} added to inventory.", "success")
            return redirect(url_for("inventory"))

        q = request.args.get("q", "").strip()
        if q:
            rows = db.execute(text("""
                SELECT * FROM inventory
                WHERE LOWER(product_name) LIKE LOWER(:q) OR LOWER(category) LIKE LOWER(:q)
                ORDER BY product_name
            """), {"q": f"%{q}%"}).mappings().all()
        else:
            rows = db.execute(text("SELECT * FROM inventory ORDER BY product_name")).mappings().all()
    except Exception as e:
        db.rollback()
        flash(str(e), "error")
        rows = db.execute(text("SELECT * FROM inventory ORDER BY product_name")).mappings().all()
    finally:
        db.close()
    return render_template("inventory.html", items=rows, search=request.args.get("q", ""))


@app.route("/sale", methods=["GET", "POST"])
@require_login
def sale():
    db = get_db()
    try:
        products = db.execute(text("SELECT id, product_name, selling_price, stock_qty FROM inventory ORDER BY product_name")).mappings().all()
        if request.method == "POST":
            product_id = int(request.form.get("product_id"))
            quantity = int(request.form.get("quantity", 0) or 0)
            payment_method = request.form.get("payment_method", "Cash")
            unit_price = float(request.form.get("unit_price", 0) or 0)
            product = db.execute(text("SELECT * FROM inventory WHERE id = :id"), {"id": product_id}).mappings().first()
            if not product:
                raise ValueError("Product not found.")
            if quantity <= 0:
                raise ValueError("Quantity must be more than zero.")
            if product["stock_qty"] < quantity:
                raise ValueError(f"Not enough stock. Available: {product['stock_qty']}")
            total_price = unit_price * quantity
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db.execute(text("""
                INSERT INTO sales (sale_date, product_id, quantity, unit_price, total_price, payment_method, attendant_name)
                VALUES (:sale_date, :product_id, :quantity, :unit_price, :total_price, :payment_method, :attendant_name)
            """), {
                "sale_date": now,
                "product_id": product_id,
                "quantity": quantity,
                "unit_price": unit_price,
                "total_price": total_price,
                "payment_method": payment_method,
                "attendant_name": session.get("full_name"),
            })
            db.execute(text("UPDATE inventory SET stock_qty = stock_qty - :qty, updated_at = :updated_at WHERE id = :id"), {
                "qty": quantity,
                "updated_at": now,
                "id": product_id,
            })
            db.commit()
            flash(f"Sale saved for {product['product_name']}.", "success")
            return redirect(url_for("sale"))
    except Exception as e:
        db.rollback()
        flash(str(e), "error")
    finally:
        db.close()
    return render_template("sale.html", products=products)


@app.route("/restock", methods=["GET", "POST"])
@require_login
def restock():
    db = get_db()
    try:
        products = db.execute(text("SELECT id, product_name, cost_price, stock_qty FROM inventory ORDER BY product_name")).mappings().all()
        if request.method == "POST":
            product_id = int(request.form.get("product_id"))
            quantity = int(request.form.get("quantity", 0) or 0)
            cost_price = float(request.form.get("cost_price", 0) or 0)
            if quantity <= 0:
                raise ValueError("Quantity must be more than zero.")
            product = db.execute(text("SELECT * FROM inventory WHERE id = :id"), {"id": product_id}).mappings().first()
            if not product:
                raise ValueError("Product not found.")
            total_cost = quantity * cost_price
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            db.execute(text("""
                INSERT INTO purchases (purchase_date, product_id, quantity, cost_price, total_cost, attendant_name)
                VALUES (:purchase_date, :product_id, :quantity, :cost_price, :total_cost, :attendant_name)
            """), {
                "purchase_date": now,
                "product_id": product_id,
                "quantity": quantity,
                "cost_price": cost_price,
                "total_cost": total_cost,
                "attendant_name": session.get("full_name"),
            })
            db.execute(text("""
                UPDATE inventory
                SET stock_qty = stock_qty + :qty, cost_price = :cost_price, updated_at = :updated_at
                WHERE id = :id
            """), {
                "qty": quantity,
                "cost_price": cost_price,
                "updated_at": now,
                "id": product_id,
            })
            db.commit()
            flash(f"Stock updated for {product['product_name']}.", "success")
            return redirect(url_for("restock"))
    except Exception as e:
        db.rollback()
        flash(str(e), "error")
    finally:
        db.close()
    return render_template("restock.html", products=products)


@app.route("/daily-report")
@require_login
def daily_report():
    report_date = request.args.get("report_date", date.today().isoformat())
    db = get_db()
    try:
        summary = db.execute(text("""
            SELECT
                COALESCE((SELECT SUM(total_price) FROM sales WHERE DATE(sale_date)=:d), 0) AS total_sales,
                COALESCE((SELECT SUM(quantity) FROM sales WHERE DATE(sale_date)=:d), 0) AS items_sold,
                COALESCE((SELECT COUNT(*) FROM sales WHERE DATE(sale_date)=:d), 0) AS sales_count,
                COALESCE((SELECT SUM(total_cost) FROM purchases WHERE DATE(purchase_date)=:d), 0) AS total_restock
        """), {"d": report_date}).mappings().first()

        sales_rows = db.execute(text("""
            SELECT s.sale_date, i.product_name, s.quantity, s.unit_price, s.total_price, s.payment_method, s.attendant_name
            FROM sales s JOIN inventory i ON i.id = s.product_id
            WHERE DATE(s.sale_date)=:d
            ORDER BY s.id DESC
        """), {"d": report_date}).mappings().all()

        restock_rows = db.execute(text("""
            SELECT p.purchase_date, i.product_name, p.quantity, p.cost_price, p.total_cost, p.attendant_name
            FROM purchases p JOIN inventory i ON i.id = p.product_id
            WHERE DATE(p.purchase_date)=:d
            ORDER BY p.id DESC
        """), {"d": report_date}).mappings().all()
    finally:
        db.close()
    return render_template("daily_report.html", report_date=report_date, summary=summary, sales_rows=sales_rows, restock_rows=restock_rows)


@app.route("/daily-report/export")
@require_login
def export_daily_report():
    report_date = request.args.get("report_date", date.today().isoformat())
    db = get_db()
    try:
        sales_rows = db.execute(text("""
            SELECT s.sale_date, i.product_name, s.quantity, s.unit_price, s.total_price, s.payment_method, s.attendant_name
            FROM sales s JOIN inventory i ON i.id = s.product_id
            WHERE DATE(s.sale_date)=:d
            ORDER BY s.id DESC
        """), {"d": report_date}).mappings().all()
        restock_rows = db.execute(text("""
            SELECT p.purchase_date, i.product_name, p.quantity, p.cost_price, p.total_cost, p.attendant_name
            FROM purchases p JOIN inventory i ON i.id = p.product_id
            WHERE DATE(p.purchase_date)=:d
            ORDER BY p.id DESC
        """), {"d": report_date}).mappings().all()
    finally:
        db.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Daily Report", report_date])
    writer.writerow([])
    writer.writerow(["Sales"])
    writer.writerow(["Date", "Product", "Qty", "Unit Price", "Total", "Payment", "Attendant"])
    for row in sales_rows:
        writer.writerow([row["sale_date"], row["product_name"], row["quantity"], row["unit_price"], row["total_price"], row["payment_method"], row["attendant_name"]])
    writer.writerow([])
    writer.writerow(["Restocks"])
    writer.writerow(["Date", "Product", "Qty", "Cost Price", "Total Cost", "Attendant"])
    for row in restock_rows:
        writer.writerow([row["purchase_date"], row["product_name"], row["quantity"], row["cost_price"], row["total_cost"], row["attendant_name"]])

    response = make_response(output.getvalue())
    response.headers["Content-Type"] = "text/csv"
    response.headers["Content-Disposition"] = f"attachment; filename=daily_report_{report_date}.csv"
    return response


@app.route("/users", methods=["GET", "POST"])
@require_role("admin")
def users():
    db = get_db()
    try:
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            full_name = request.form.get("full_name", "").strip()
            password = request.form.get("password", "").strip()
            role = request.form.get("role", "attendant").strip()
            if not username or not full_name or not password:
                raise ValueError("Username, full name and password are required.")
            db.execute(text("""
                INSERT INTO users (username, full_name, password_hash, role, is_active, created_at)
                VALUES (:username, :full_name, :password_hash, :role, 1, :created_at)
            """), {
                "username": username,
                "full_name": full_name,
                "password_hash": generate_password_hash(password),
                "role": role,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            db.commit()
            flash(f"{full_name} added successfully.", "success")
            return redirect(url_for("users"))

        all_users = db.execute(text("SELECT id, username, full_name, role, is_active, created_at FROM users ORDER BY id ASC")).mappings().all()
    except Exception as e:
        db.rollback()
        flash(str(e), "error")
        all_users = db.execute(text("SELECT id, username, full_name, role, is_active, created_at FROM users ORDER BY id ASC")).mappings().all()
    finally:
        db.close()
    return render_template("users.html", users=all_users)


@app.route("/health")
def health():
    return {"ok": True}


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=True)
